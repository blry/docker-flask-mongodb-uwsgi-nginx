from openpyxl import load_workbook
from project import utils, db


def main():
    parsed_files_col, parsed_records_col = db.working_collections()
    parsed_files = list(parsed_files_col.find())

    all_files = utils.all_files('xlsx')
    print('Found %d files in xlsx dir:' % len(all_files))

    for file in all_files:
        if file not in parsed_files:
            print('%s is processing' % file['path'])
            records = parse_records(file['path'])
            parsed_records_col.insert_many(records)
            parsed_files_col.insert_one(file)
            print('Done. Added %d records' % len(records))
        else:
            print('%s is already processed' % file['path'])

    print('Finished!')


def parse_records(filepath):
    records = []
    workbook = load_workbook(filepath, True)

    for worksheet in workbook.worksheets:
        week_start, week_end = utils.week_dates(worksheet.cell(row=3, column=1).value)
        record = {'week_start': week_start, 'week_end': week_end}
        row = 6
        while True:
            row += 1
            name = worksheet.cell(row=row, column=1).value or worksheet.cell(row=row, column=2).value
            if name is None or name == '':
                break
            if 'Total' in name:
                continue

            record[name.replace('.', '').replace('$', '')] = int(worksheet.cell(row = row, column=3).value)

        records.append(record)

    return records


if __name__ == '__main__':
    main()
